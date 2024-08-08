import logging
import os
import re
import time
import zipfile
import configparser
from io import StringIO
from typing import List

import pandas as pd
import xlrd
from openpyxl import load_workbook
from IPython.display import display

from enums import FileExtension


class ComparisonExcel:
    """Class for comparing Excel and INI files."""

    def __init__(self):
        # Configure logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def get_active_sheet_name(self, file_path: str) -> str:
        """
        Get the name of the active sheet in an Excel file.
        
        Args:
            file_path (str): Path to the Excel file.

        Returns:
            str: Name of the active sheet.
        """
        workbook = load_workbook(file_path)
        return workbook.active.title

    def read_excel_file(self, file_path: str, sheet_name: str, init_column: int, row_value: int) -> pd.DataFrame:
        """
        Read an Excel file and extract parameter names and values.

        Args:
            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet to read.
            init_column (int): Initial column index (1-based).
            row_value (int): Row index (1-based) for values.

        Returns:
            pd.DataFrame: DataFrame containing parameter names and values.
        """
        df = pd.read_excel(file_path, header=None, sheet_name=sheet_name)
        idx_col = init_column - 1
        parameter_names = df.iloc[0, idx_col:].tolist()
        parameter_values = df.iloc[row_value - 1, idx_col:].tolist()
        return pd.DataFrame({
            "Parameter_Name": parameter_names,
            "Parameter_Value": parameter_values
        })

    def compare_excel_files(self, reference_file: str, downloaded_file: str, part_number: str) -> list:
        """
        Compare the contents of two Excel files.

        Args:
            reference_file (str): Path to the reference Excel file.
            downloaded_file (str): Path to the downloaded Excel file.
            part_number (str): Part number for reporting.

        Returns:
            list: Comparison results and messages.
        """
        reference_sheet = self.get_active_sheet_name(reference_file)
        df_reference = self.read_excel_file(reference_file, reference_sheet, init_column=3, row_value=17)

        downloaded_excel_file = pd.ExcelFile(downloaded_file)
        pattern = r"Z\d{6}"
        downloaded_sheet = next((sheet for sheet in downloaded_excel_file.sheet_names if re.search(pattern, sheet)), None)
        df_downloaded = self.read_excel_file(downloaded_file, downloaded_sheet, init_column=3, row_value=17)

        missing_param_names = self.find_missing_params(df_reference, df_downloaded)
        msg_pdf = self.format_missing_params_message(missing_param_names)

        comparison_result = self.perform_comparison(df_reference, df_downloaded)
        msg_results = f"\nResults for {part_number}:\n"
        print(f"\033[1m{msg_results}\033[0m")
        msg_pdf += msg_results

        styled_df = self.style_comparison_result(comparison_result)
        display(styled_df)
        return [{"msg": msg_pdf, "df": comparison_result}]

    def find_missing_params(self, df_reference: pd.DataFrame, df_downloaded: pd.DataFrame) -> set:
        """
        Find parameters present in the reference DataFrame but missing in the downloaded DataFrame.

        Args:
            df_reference (pd.DataFrame): Reference DataFrame.
            df_downloaded (pd.DataFrame): Downloaded DataFrame.

        Returns:
            set: Set of missing parameter names.
        """
        reference_param_names = set(df_reference["Parameter_Name"].unique())
        downloaded_param_names = set(df_downloaded["Parameter_Name"].unique())
        return reference_param_names - downloaded_param_names

    def format_missing_params_message(self, missing_param_names: set) -> str:
        """
        Format the message for missing parameters.

        Args:
            missing_param_names (set): Set of missing parameter names.

        Returns:
            str: Formatted message for missing parameters.
        """
        if not missing_param_names:
            return ""
        msg_pdf = "\nThe following parameters from the reference file are not present in the downloaded file:\n"
        print(f"\033[1m\033[48;5;208m{msg_pdf}\033[0m")
        for param_name in missing_param_names:
            param_msg = f"- {param_name}"
            print(f"\033[1;38;2;255;0;0m{param_msg}\033[0m")
            msg_pdf += f"{param_msg}\n"
        return msg_pdf

    def perform_comparison(self, df_reference: pd.DataFrame, df_downloaded: pd.DataFrame) -> pd.DataFrame:
        """
        Perform the comparison between reference and downloaded DataFrames.

        Args:
            df_reference (pd.DataFrame): Reference DataFrame.
            df_downloaded (pd.DataFrame): Downloaded DataFrame.

        Returns:
            pd.DataFrame: DataFrame with comparison results.
        """
        df_reference.rename(columns={"Parameter_Name": "Parameter_Name", "Parameter_Value": "Expected_Value"}, inplace=True)
        df_downloaded.rename(columns={"Parameter_Name": "Parameter_Name", "Parameter_Value": "Your_INI_Value"}, inplace=True)

        comparison_result = pd.merge(df_reference, df_downloaded, on="Parameter_Name", how="inner")
        return comparison_result.fillna("")

    def style_comparison_result(self, comparison_result: pd.DataFrame) -> pd.DataFrame.style:
        """
        Apply styling to the comparison result DataFrame.

        Args:
            comparison_result (pd.DataFrame): DataFrame with comparison results.

        Returns:
            pd.DataFrame.style: Styled DataFrame.
        """
        def highlight_mismatched_values(val):
            color = "red" if val["Expected_Value"] != val["Your_INI_Value"] else "black"
            font_weight = "bold" if val["Expected_Value"] != val["Your_INI_Value"] else "normal"
            return [f"color: {color}; font-weight: {font_weight}" for _ in val]

        return comparison_result.style.apply(
            highlight_mismatched_values,
            axis=1,
            subset=["Parameter_Name", "Expected_Value", "Your_INI_Value"]
        ).set_table_styles(
            [
                {"selector": "th", "props": [("text-align", "left"), ("font-weight", "bold")]},
                {"selector": "td:nth-child(2)", "props": [("width", "250px"), ("text-align", "left"), ("word-break", "break-word")]},
                {"selector": "td:nth-child(3)", "props": [("width", "400px"), ("text-align", "left"), ("word-break", "break-word")]},
                {"selector": "td:nth-child(4)", "props": [("width", "400px"), ("text-align", "left"), ("word-break", "break-word")]},
            ]
        )

    def read_ini_file(self, ini_file_path: str, titles: list, part_number: str) -> pd.DataFrame:
        """
        Read and parse an INI file to extract parameters.

        Args:
            ini_file_path (str): Path to the INI file.
            titles (list): List of titles for parameter extraction.
            part_number (str): Part number for specific parameters.

        Returns:
            pd.DataFrame: DataFrame containing parameters and their values.
        """
        config = configparser.ConfigParser()
        config.read(ini_file_path)

        data = []
        for section in config.sections():
            if section.startswith(('PARAM', 'DTCPARAM')):
                parameter_name = config.get(section, 'PARAMETERNAME', fallback=None).strip('"')
                parameter_value = self.extract_ini_parameter_value(parameter_name, config, section, titles, part_number)
                if parameter_name:
                    data.append({'Parameter_Name': parameter_name, 'Parameter_Value': parameter_value})

        return pd.DataFrame(data)

    def extract_ini_parameter_value(self, parameter_name: str, config: configparser.ConfigParser, section: str, titles: List[List[str]], part_number: str) -> str:
        """
        Extract the value for a given INI parameter based on its name.

        Args:
            parameter_name (str): The name of the parameter.
            config (configparser.ConfigParser): ConfigParser object for reading INI file.
            section (str): INI section name.
            titles (list): List of titles for parameter extraction.
            part_number (str): Part number for specific parameters.

        Returns:
            str: Value for the parameter.
        """
        # Use helper functions to find values
        if parameter_name in ["BOOT_SW_ID", "SS_APP_SW_ID"]:
            return self._extract_boot_or_app_id(parameter_name, titles)
        
        if parameter_name in ["APP_SW_ID", "SS_ECU_SW_NUMBER"]:
            return self._find_value_in_titles('Software', titles=titles, exclude_boot=True)
        
        if parameter_name == "APP_DATA_ID":
            return self._find_value_in_titles('Dataset', titles=titles)
        
        if parameter_name == "VEH_MAN_ECU_HW_NUMBER":
            return part_number
        
        # Default case: read from INI file
        return config.get(section, 'PARAMETERVALUE', fallback="Undefined").strip('"')
    
    def _extract_boot_or_app_id(self, parameter_name: str, titles: List[List[str]]) -> str:
        """
        Extracts the BOOT_SW_ID or SS_APP_SW_ID based on titles.

        Args:
            parameter_name (str): The name of the parameter.
            titles (List[List[str]]): List of titles for parameter extraction.

        Returns:
            str: Value for the parameter.
        """
        na_pattern = re.compile(r'NA(\d+)')
        results = [
            na_match.group(1) for item in titles
            if 'boot' in item[2].lower() and (na_match := na_pattern.search(item[2]))
        ]
        if parameter_name == "BOOT_SW_ID":
            return "NA" + results[0] if results else "Undefined"
        if parameter_name == "SS_APP_SW_ID":
            return "DCCANA" + results[0] if results else "Undefined"
        return "Undefined"

    def _find_value_in_titles(self, keyword: str, titles: List[List[str]], exclude_boot: bool = False) -> str:
        """
        Filters titles based on keyword and optional exclusion of 'boot'.

        Args:
            keyword (str): Keyword to match in titles.
            titles (List[List[str]]): List of titles for parameter extraction.            
            exclude_boot (bool): Whether to exclude titles containing 'boot'.

        Returns:
            str: Value from the first matching title or "Undefined".
        """
        for item in titles:
            if keyword in item[0] and (not exclude_boot or 'boot' not in item[2].lower()):
                return item[3]
        return "Undefined"

    def get_ini_from_zip(self, zip_file_path: str) -> pd.DataFrame:
        """
        Extract and parse INI files from a ZIP archive.

        Args:
            zip_file_path (str): Path to the ZIP file.

        Returns:
            pd.DataFrame: DataFrame containing parameters and values extracted from INI files.
        """
        data = []
        try:
            time.sleep(3)  # To ensure the file is fully accessible
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                for file_name in zip_ref.namelist():
                    if file_name.endswith(FileExtension.INI.value): 
                        try:
                            with zip_ref.open(file_name) as file:
                                ini_content = file.read().decode('utf-8')
                            ini_file = StringIO(ini_content)
                            config = configparser.ConfigParser()
                            config.read_file(ini_file)
                            for section in config.sections():
                                if section.startswith(('PARAM', 'DTCPARAM')):
                                    parameter_name = config.get(section, 'PARAMETERNAME', fallback=None).strip('"')
                                    parameter_value = config.get(section, 'PARAMETERVALUE', fallback=None).strip('"')
                                    if parameter_name:
                                        data.append({'Parameter_Name': parameter_name, 'Parameter_Value': parameter_value})
                        except Exception as e:
                            self.logger.error(f"Error processing INI file '{file_name}': {e}")
        except zipfile.BadZipFile:
            self.logger.error(f"The file '{zip_file_path}' is not a valid ZIP file.")
        except Exception as e:
            self.logger.error(f"An error occurred while reading the ZIP file '{zip_file_path}': {e}")

        return pd.DataFrame(data)

    def compare_excel_ini_files(self, reference_file: str, downloaded_file: str, part_number: str, titles: list) -> list:
        """
        Compare the contents of an INI file from a ZIP archive with an Excel file.

        Args:
            reference_file (str): Path to the reference Excel file.
            downloaded_file (str): Path to the ZIP file containing INI files.
            part_number (str): Part number for reporting.
            titles (list): List of titles for parameter extraction.

        Returns:
            list: List containing comparison results and messages.
        """
        df_reference = self.read_ini_file(reference_file, titles, part_number)
        df_downloaded = self.get_ini_from_zip(downloaded_file)

        reference_param_names = set(df_reference["Parameter_Name"])
        downloaded_param_names = set(df_downloaded["Parameter_Name"])
        missing_param_names = reference_param_names - downloaded_param_names

        msg_pdf = self.format_missing_params_message(missing_param_names)

        df_reference.rename(columns={"Parameter_Name": "Parameter_Name", "Parameter_Value": "Expected_Value"}, inplace=True)
        df_downloaded.rename(columns={"Parameter_Name": "Parameter_Name", "Parameter_Value": "Your_INI_Value"}, inplace=True)

        comparison_result = pd.merge(df_reference, df_downloaded, on="Parameter_Name", how="inner").fillna("")

        msg_results = f"\nResults for {part_number}:\n"
        print(f"\033[1m{msg_results}\033[0m")
        msg_pdf += msg_results

        styled_df = self.style_comparison_result(comparison_result)
        display(styled_df)

        return [{"msg": msg_pdf, "df": comparison_result}]

    def comparison(self, file_name_path: str, reference_file: str, part_number: str) -> list:
        """
        Run the comparison for Excel files.

        Args:
            file_name_path (str): Path to the downloaded file.
            reference_file (str): Path to the reference Excel file.
            part_number (str): Part number for reporting.

        Returns:
            list: Comparison results.
        """
        try:
            valid_extensions = {FileExtension.XLS.value, FileExtension.XLSX.value}
            if any(file_name_path.endswith(ext) for ext in valid_extensions):
                return self.compare_excel_files(reference_file, file_name_path, part_number)
            else:
                print("The file was not processed, since the extension is not .xlsx or .xls")
                return []
        except Exception as e:
            self.logger.error("An error occurred: %s", e)
            return []

    def comparison_ini(self, file_name_path: str, reference_file: str, part_number: str, titles: list) -> list:
        """
        Run the comparison for INI files from a ZIP archive.

        Args:
            file_name_path (str): Path to the ZIP file containing INI files.
            reference_file (str): Path to the reference Excel file.
            part_number (str): Part number for reporting.
            titles (list): List of titles for parameter extraction.

        Returns:
            list: Comparison results.
        """
        try:
            if file_name_path.endswith(FileExtension.INI.value): 
                return self.compare_excel_ini_files(reference_file, file_name_path, part_number, titles)
            else:
                print("The file was not processed, since the extension is not .zip")
                return []
        except Exception as e:
            self.logger.error("An error occurred: %s", e)
            return []

    def has_red_font(self, xf, book) -> bool:
        """
        Check if a cell has red font.

        Args:
            xf: Cell's XF record.
            book: Workbook object.

        Returns:
            bool: True if the cell's font is red, False otherwise.
        """
        if xf.font_index is not None:
            font = book.font_list[xf.font_index]
            font_color_rgb = book.colour_map.get(font.colour_index)
            return font_color_rgb == (255, 0, 0)
        return False

    def has_table_border(self, xf) -> bool:
        """
        Check if a cell has a table border.

        Args:
            xf: Cell's XF record.

        Returns:
            bool: True if the cell has borders, False otherwise.
        """
        border_style = xf.border
        return all([
            border_style.top_line_style,
            border_style.bottom_line_style,
            border_style.left_line_style,
            border_style.right_line_style
        ])

    def is_bold(self, xf, workbook) -> bool:
        """
        Check if a cell's font is bold.

        Args:
            xf: Cell's XF record.
            workbook: Workbook object.

        Returns:
            bool: True if the cell's font is bold, False otherwise.
        """
        font = workbook.font_list[xf.font_index]
        return font.bold == 1

    def has_red_font_and_table_border(self, cell, workbook) -> bool:
        """
        Check if a cell has red font and a table border.

        Args:
            cell: Cell object.
            workbook: Workbook object.

        Returns:
            bool: True if the cell has red font and a table border, False otherwise.
        """
        xf_index = cell.xf_index
        xf = workbook.xf_list[xf_index]
        return self.has_red_font(xf, workbook) and self.has_table_border(xf)

    def has_borders_and_bold(self, cell, workbook) -> bool:
        """
        Check if a cell has borders and bold text.

        Args:
            cell: Cell object.
            workbook: Workbook object.

        Returns:
            bool: True if the cell has borders and bold text, False otherwise.
        """
        xf = workbook.xf_list[cell.xf_index]
        return self.is_bold(xf, workbook) and self.has_table_border(xf)

    def find_table_start(self, sheet, workbook) -> int:
        """
        Find the starting row index of the table by looking for the first row with cells that have borders and bold text.

        Args:
            sheet: xlrd sheet object.
            workbook: xlrd workbook object.

        Returns:
            int: Row index where the table starts, or None if not found.
        """
        for i in range(sheet.nrows):
            cell = sheet.cell(i, 0)
            xf = workbook.xf_list[cell.xf_index]
            if (self.is_bold(xf, workbook) and self.has_table_border(xf)) or (
                self.is_bold(xf, workbook) and xf.background.fill_pattern
            ):
                return i
        return None

    def clean_header(self, header: str) -> str:
        """
        Clean the header string by replacing new lines with spaces and removing extra spaces.

        Args:
            header (str): Header string to clean.

        Returns:
            str: Cleaned header string.
        """
        return " ".join(header.replace("\n", " ").split())

    def extract_info_schedule(self, file_path: str) -> dict:
        """
        Extract information from the 'SCHEDULE' sheet in an Excel file.

        Args:
            file_path (str): Path to the Excel file.

        Returns:
            dict: Dictionary containing row information and headers, or an empty list if not found.
        """
        if file_path.endswith(FileExtension.XLS.value):
            workbook = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = workbook.sheet_by_name("SCHEDULE")

            table_start_row_idx = self.find_table_start(sheet, workbook)

            if table_start_row_idx is not None:
                headers_row = [sheet.cell_value(table_start_row_idx, col) for col in range(sheet.ncols)]
                headers = [self.clean_header(header) for header in headers_row]

                red_font_rows = [
                    i
                    for i in range(table_start_row_idx + 1, sheet.nrows)
                    if self.has_red_font_and_table_border(sheet.cell(i, 0), workbook) or (
                        self.has_red_font(
                            workbook.xf_list[sheet.cell(i, 0).xf_index], workbook
                        ) and re.search(r"^K\d+", sheet.cell(i, 0).value)
                    )
                ]
                if not red_font_rows:
                    print(f"\033[1m\033[48;5;208mNo Part number was found to be processed. \nNote: To process a part number it should be with red font (starting from column 0) in the excel file.\n\033[0m")
                    return None
                row_info = [
                    [
                        sheet.cell_value(row, col).strip()
                        if isinstance(sheet.cell_value(row, col), str)
                        else sheet.cell_value(row, col)
                        for col in range(sheet.ncols)
                    ]
                    for row in red_font_rows
                ]
                return {"row_info": row_info, "headers": headers}
            else:
                print(f"\033[1m\033[48;5;208mNo headers were found.\n\033[0m")
                return []
        else:
            print(f"The extension {os.path.splitext(file_path)[1]} cannot be processed. Expected extension is '.xls'")
            return []

    def get_info_red_row(self, row_info: list, headers: list) -> dict:
        """
        Map headers to their corresponding values from the red font rows.

        Args:
            row_info (list): List of row information.
            headers (list): List of headers.

        Returns:
            dict: Dictionary mapping cleaned header names to their values.
        """
        headers_dict = {
            header: value for header, value in zip(headers, row_info) if header.strip()
        }
        field_mapping = {
            "PART NUMBER": [
                "BENDIX PART NUMBER",
                "REMANUFACTURED PART NUMBER",
                "BENDIX AFTER MARKET PART NUMBER",
            ],
            "CONFIG INI DATA SET": [
                "DATA SET PART NUMBER / CONFIGURATION",
                "CONFIGURATION",
                "DATASET PART NUMBER",
            ],
            "CUSTOMER PART NUMBER": [
                "CUSTOMER PART NUMBER / CUSTOMER SOFTWARE PART NUMBER"
            ],
            "PART NUMBER LABEL": ["LABEL"],
        }
        updated_headers = {
            field: value
            for key, value in headers_dict.items()
            for field, aliases in field_mapping.items()
            if key in aliases
        }
        updated_headers.update({
            key: value
            for key, value in headers_dict.items()
            if key not in updated_headers
        })
        return updated_headers
